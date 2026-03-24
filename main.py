import asyncio
import logging
import sys
import os
import requests
import json
from datetime import datetime, timedelta, date as date_cls, time as time_cls
from typing import Dict, Optional, List, Set
from telethon import TelegramClient, events
from telethon.errors import FloodWaitError
from telethon.sessions import StringSession
from telethon.tl.functions.bots import SetBotCommandsRequest
from telethon.tl.types import BotCommand, BotCommandScopeDefault
from aiohttp import web

from config import (
    API_ID, API_HASH, BOT_TOKEN, ADMIN_ID,
    TELEGRAM_SESSION, PORT, is_admin
)
from database import (
    init_db, save_game, get_games_by_date,
    get_dates_available, get_stats_by_date, count_games_for_date,
    get_last_saved_game_num, get_comptage_today, get_global_total,
    create_public_view, get_db_size, init_documentation,
    search_games_by_filter, stats_from_rows, FILTRES_DISPONIBLES,
    compare_dates, get_all_games_for_export,
    search_games_multi_filter, row_matches_filter
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

if not API_ID or API_ID == 0:
    logger.error("API_ID manquant"); sys.exit(1)
if not API_HASH:
    logger.error("API_HASH manquant"); sys.exit(1)
if not BOT_TOKEN:
    logger.error("BOT_TOKEN manquant"); sys.exit(1)

client = None

# Etat de conversation: {sender_id: "awaiting_date"}
active_conversations: set = set()  # évite les doubles conversations simultanées

# Jeux déjà enregistrés dans cette session (évite les doublons)
saved_games_session: Set[int] = set()


# ============================================================================
# UTILITAIRES TEMPS
# ============================================================================

def get_local_time() -> datetime:
    return datetime.utcnow()  # Côte d'Ivoire = GMT (UTC+0)


def game_num_to_time(num: int) -> time_cls:
    """Calcule l'heure GMT estimée d'un jeu selon son numéro (1=00:00, 1440=23:59)."""
    minutes = (num - 1) % 1440
    return time_cls(hour=minutes // 60, minute=minutes % 60)


def current_game_date() -> date_cls:
    """La date courante GMT (Côte d'Ivoire)."""
    return get_local_time().date()


def determine_game_date(game_number: int) -> date_cls:
    """
    Détermine la vraie date d'un jeu selon son numéro.
    Les jeux vont de #1 (00:00 GMT) à #1440 (23:59 GMT).
    Si le bot tourne après minuit et l'API retourne encore des jeux
    à numéros élevés (ex: #1421 à 00:40 GMT), ce sont les derniers
    jeux du jour précédent.
    """
    now = get_local_time()
    today = now.date()
    yesterday = today - timedelta(days=1)
    # Numéro attendu maintenant
    expected_num = now.hour * 60 + now.minute + 1
    # Si le jeu a un numéro > 12h d'avance sur l'heure actuelle → jour précédent
    if game_number > expected_num + 720:
        return yesterday
    return today


# ============================================================================
# API 1xBet - RECUPERATION DES PARTIES
# ============================================================================

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json",
    "Referer": "https://1xbet.com/",
}

SUIT_MAP   = {0: "Pique", 1: "Trefle", 2: "Carreau", 3: "Coeur"}
RANK_MAP   = {1:"As",2:"2",3:"3",4:"4",5:"5",6:"6",7:"7",8:"8",9:"9",10:"10",11:"Valet",12:"Dame",13:"Roi"}
RANK_SHORT = {1:"A",2:"2",3:"3",4:"4",5:"5",6:"6",7:"7",8:"8",9:"9",10:"10",11:"J",12:"Q",13:"K"}
ENS_EMOJI  = {0:"♠",1:"♣",2:"♦",3:"♥"}


def parse_cards(sc_s_list: list):
    player_cards, banker_cards = [], []
    for entry in sc_s_list:
        key = entry.get("Key", "")
        try:
            cards = json.loads(entry.get("Value", "[]"))
        except Exception:
            cards = []
        if key == "P":
            player_cards = cards
        elif key == "B":
            banker_cards = cards
    return player_cards, banker_cards


def parse_winner(sc_s_list: list) -> Optional[str]:
    for entry in sc_s_list:
        if entry.get("Key") == "S":
            v = entry.get("Value", "")
            if v in ("Win1",):   return "Player"
            if v in ("Win2",):   return "Banker"
            if v in ("Tie", "Draw"): return "Tie"
    return None


def is_finished(game: dict) -> bool:
    sc = game.get("SC", {})
    s_list = sc.get("S", [])
    winner = parse_winner(s_list)
    cards_p, cards_b = parse_cards(s_list)
    return winner is not None and len(cards_p) >= 2 and len(cards_b) >= 2


def fetch_live_games() -> List[dict]:
    try:
        r = requests.get(
            "https://1xbet.com/service-api/LiveFeed/GetSportsShortZip",
            params={"sports": 236, "champs": 2050671, "lng": "en",
                    "gr": 285, "country": 96, "virtualSports": "true",
                    "groupChamps": "true"},
            headers=HEADERS, timeout=20
        )
        data = r.json()
    except Exception as e:
        logger.error(f"fetch_live_games error: {e}")
        return []

    results = []
    for sport in data.get("Value", []):
        for champ in sport.get("L", []):
            for game in champ.get("G", []):
                if "DI" not in game:
                    continue
                game_number = int(game["DI"])
                sc = game.get("SC", {})
                sc_s = sc.get("S", [])
                player_raw, banker_raw = parse_cards(sc_s)
                winner = parse_winner(sc_s)
                finished = is_finished(game)

                def fmt(cards):
                    return [{"suit": c.get("S", -1), "rank": c.get("R", 0)} for c in cards]

                results.append({
                    "game_number":   game_number,
                    "player_cards":  fmt(player_raw),
                    "banker_cards":  fmt(banker_raw),
                    "winner":        winner,
                    "is_finished":   finished,
                })
    return results


# ============================================================================
# BOUCLE DE CAPTURE AUTOMATIQUE
# ============================================================================

async def polling_loop():
    """Capture chaque partie terminée et la sauvegarde en DB."""
    logger.info("📡 Boucle de capture automatique démarrée")
    while True:
        try:
            loop = asyncio.get_event_loop()
            games = await loop.run_in_executor(None, fetch_live_games)
            saved_count = 0

            for game in games:
                gnum = game["game_number"]
                if not game["is_finished"]:
                    continue
                if gnum in saved_games_session:
                    continue

                game_date = determine_game_date(gnum)
                heure = game_num_to_time(gnum)
                was_saved = await loop.run_in_executor(
                    None, save_game, game, game_date, heure
                )
                if was_saved:
                    saved_games_session.add(gnum)
                    saved_count += 1
                    logger.info(f"💾 Jeu #{gnum} sauvegardé | {game['winner']} | {game_date}")
                else:
                    saved_games_session.add(gnum)

            if saved_count > 0:
                logger.info(f"✅ {saved_count} nouveau(x) jeu(x) enregistré(s)")

        except Exception as e:
            logger.error(f"polling_loop error: {e}")

        await asyncio.sleep(8)


# ============================================================================
# GENERATION DU PDF
# ============================================================================

def safe_text(text: str) -> str:
    rep = {'♠':'P','♣':'T','♦':'C','♥':'Co','é':'e','è':'e','ê':'e','à':'a',
           'ù':'u','â':'a','ô':'o','î':'i','ç':'c','É':'E','È':'E','À':'A',
           'Ê':'E','Ù':'U','Â':'A','Ô':'O','Î':'I','Ç':'C','–':'-','—':'-',
           '\u2026':'...', '\u201c':'"', '\u201d':'"'}
    for ch, r in rep.items():
        text = text.replace(ch, r)
    return text.encode('latin-1', errors='replace').decode('latin-1')


def generate_pdf_from_db(path: str, rows: List[dict], target_date: str, stats: dict,
                         filtre_label: str = "") -> None:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    C_DARK   = (20, 30, 60)
    C_HEAD   = (40, 60, 110)
    C_WHITE  = (255, 255, 255)
    C_LIGHT  = (235, 240, 255)
    C_GREEN  = (20, 110, 50)
    C_RED    = (170, 20, 20)
    C_BLUE   = (20, 70, 180)
    C_GOLD   = (150, 100, 0)
    C_GRAY   = (130, 130, 130)
    C_ODD    = (250, 250, 255)

    total    = int(stats.get("total", 0) or 0)
    jw       = int(stats.get("joueur_wins", 0) or 0)
    bw       = int(stats.get("banquier_wins", 0) or 0)
    egal     = int(stats.get("egalites", 0) or 0)
    naturels = int(stats.get("naturels", 0) or 0)
    paires_j = int(stats.get("paires_joueur", 0) or 0)
    paires_b = int(stats.get("paires_banquier", 0) or 0)
    j4c      = int(stats.get("jeux_4cartes", 0) or 0)
    j5c      = int(stats.get("jeux_5cartes", 0) or 0)
    j6c      = int(stats.get("jeux_6cartes", 0) or 0)
    now_str  = get_local_time().strftime("%d/%m/%Y %H:%M")

    def pct(v, t):
        return f"{round(v/t*100,1)}%" if t > 0 else "0%"

    def bar(v, t, w=22):
        f = round(v/t*w) if t > 0 else 0
        return "[" + "#"*f + "-"*(w-f) + "]"

    FONT_TTF  = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
    FONT_TTFB = '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'

    def fmt_rang(rang):
        if rang is None:
            return "-"
        return {"14": "A", "11": "J", "12": "Q", "13": "K"}.get(str(rang), str(rang))

    def fmt_card(rang, ens, val):
        """Formate une carte avec symbole Unicode : ex 5♠(5), A♦(1), K♣(0)."""
        if rang is None:
            return "-"
        r = fmt_rang(rang)
        s = str(ens) if ens else ""   # conserve ♠ ♣ ♦ ♥ directement
        v = str(val) if val is not None else "?"
        return f"{r}{s}({v})"

    # ════════════════════════════════════════════════════════════════════════
    # PAGE(S) 1 — Tableau classique (paysage A4 : 297 x 210 mm)
    # Police DejaVuSans (Unicode) pour ♠ ♣ ♦ ♥ et ✓
    # Ordre des colonnes demandé par l'utilisateur :
    #   Date · Heure · Numéro · Cartes Joueur · Cartes Banquier ·
    #   Victoire J/B · Nul · NbCartes · Pair/Impair ·
    #   Total J · +6.5J · -4.5J · Total B · +6.5B · -4.5B
    # ════════════════════════════════════════════════════════════════════════
    pdf = FPDF(orientation='L', format='A4')
    pdf.add_font('DejaVu',  '',  FONT_TTF)
    pdf.add_font('DejaVu',  'B', FONT_TTFB)
    pdf.set_auto_page_break(auto=False)

    # ── Colonnes dans l'ordre demandé ────────────────────────────────────────
    # (groupe, label_ligne1, label_ligne2, largeur_mm, clé_data)
    COLS = [
        # Infos générales
        ("",           "Date",        "",            14, "date"),
        ("",           "Heure",       "",            13, "heure"),
        ("",           "Numéro",      "",            12, "numero"),
        # Cartes Joueur
        ("JOUEUR",     "Carte 1",     "",            18, "jc1"),
        ("JOUEUR",     "Carte 2",     "",            18, "jc2"),
        ("JOUEUR",     "Carte 3",     "",            15, "jc3"),
        # Cartes Banquier
        ("BANQUIER",   "Carte 1",     "",            18, "bc1"),
        ("BANQUIER",   "Carte 2",     "",            18, "bc2"),
        ("BANQUIER",   "Carte 3",     "",            15, "bc3"),
        # Résultats
        ("RÉSULTAT",   "Victoire",    "Joueur",      13, "vict_j"),
        ("RÉSULTAT",   "Victoire",    "Banquier",    13, "vict_b"),
        ("RÉSULTAT",   "Match",       "Nul",         11, "egalite"),
        ("RÉSULTAT",   "Nb",          "Cartes",      12, "nb_cartes"),
        # Analyse
        ("ANALYSE",    "Pair /",      "Impair",      11, "pi_num"),
        ("ANALYSE",    "Total",       "Joueur",      11, "pts_j"),
        ("ANALYSE",    "Plus de",     "6.5 J",       10, "sup65_j"),
        ("ANALYSE",    "Moins de",    "4.5 J",       10, "inf45_j"),
        ("ANALYSE",    "Total",       "Banquier",    11, "pts_b"),
        ("ANALYSE",    "Plus de",     "6.5 B",       10, "sup65_b"),
        ("ANALYSE",    "Moins de",    "4.5 B",       10, "inf45_b"),
    ]
    # Largeur de chaque groupe
    GRP_ORDER  = ["", "JOUEUR", "BANQUIER", "RÉSULTAT", "ANALYSE"]
    GRP_WIDTHS: Dict[str, int] = {}
    for col in COLS:
        g = col[0]; GRP_WIDTHS[g] = GRP_WIDTHS.get(g, 0) + col[3]

    ROW_H  = 6
    HDR1_H = 7
    HDR2_H = 12
    ML     = 10

    GRP_COLORS = {
        "":          (160, 170, 195),
        "JOUEUR":    (20, 70, 180),
        "BANQUIER":  (160, 20, 20),
        "RÉSULTAT":  (20, 100, 50),
        "ANALYSE":   (110, 60, 140),
    }

    def dv(bold=False, size=7):
        pdf.set_font('DejaVu', 'B' if bold else '', size)

    def draw_page_header():
        pdf.set_fill_color(*C_DARK)
        pdf.rect(0, 0, 297, 26, 'F')
        pdf.set_text_color(*C_WHITE)
        dv(True, 15)
        pdf.set_xy(ML, 3)
        titre = f"RECHERCHE [{filtre_label}] — BACCARAT 1xBet" if filtre_label else "HISTORIQUE BACCARAT 1xBet"
        pdf.cell(277, 10, titre, align='C',
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        dv(False, 8)
        pdf.set_xy(ML, 14)
        pdf.cell(277, 8,
                 f"Date : {target_date}  |  {total} parties  |  {now_str} GMT  |  "
                 f"Joueur : {jw} ({pct(jw,total)})  "
                 f"Banquier : {bw} ({pct(bw,total)})  "
                 f"Egalite : {egal} ({pct(egal,total)})",
                 align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    def draw_col_headers():
        y = pdf.get_y()
        # Niveau 1 : groupes colorés
        pdf.set_xy(ML, y)
        for g in GRP_ORDER:
            w = GRP_WIDTHS.get(g, 0)
            if not w:
                continue
            pdf.set_fill_color(*GRP_COLORS[g])
            pdf.set_text_color(*C_WHITE)
            dv(True, 8)
            pdf.cell(w, HDR1_H, g, border=1, fill=True, align='C')
        pdf.ln()

        # Niveau 2 : labels colonnes sur 2 lignes
        y2 = pdf.get_y()
        x  = ML
        for _, l1, l2, w, _ in COLS:
            pdf.set_fill_color(*C_HEAD)
            pdf.rect(x, y2, w, HDR2_H, 'FD')
            pdf.set_text_color(*C_WHITE)
            if l2:
                dv(True, 5)
                pdf.set_xy(x, y2 + 1.5)
                pdf.cell(w, 4, l1, align='C')
                pdf.set_xy(x, y2 + 6)
                pdf.cell(w, 4, l2, align='C')
            else:
                dv(True, 6)
                pdf.set_xy(x, y2 + 3.5)
                pdf.cell(w, 5, l1, align='C')
            x += w
        pdf.set_xy(ML, y2 + HDR2_H)
        pdf.ln()

    def draw_data_row(idx, row):
        bg = C_LIGHT if idx % 2 == 0 else C_ODD
        gagnant  = row.get("gagnant", "?")
        num_jeu  = int(row.get("numero_jeu", 0) or 0)
        pts_j    = int(row.get("joueur_points",   0) or 0)
        pts_b    = int(row.get("banquier_points", 0) or 0)
        pj_nb    = int(row.get("joueur_total_cartes")  or 2)
        pb_nb    = int(row.get("banquier_total_cartes") or 2)
        date_str = row.get("date_jeu").strftime("%d/%m/%Y") if row.get("date_jeu") else target_date
        heure    = str(row.get("heure", ""))[:5] if row.get("heure") else "--:--"

        data = {
            "date":      date_str,
            "heure":     heure,
            "numero":    str(num_jeu),
            "jc1":  fmt_card(row.get("joueur_carte1_rang"), row.get("joueur_carte1_ens"), row.get("joueur_carte1_val")),
            "jc2":  fmt_card(row.get("joueur_carte2_rang"), row.get("joueur_carte2_ens"), row.get("joueur_carte2_val")),
            "jc3":  fmt_card(row.get("joueur_carte3_rang"), row.get("joueur_carte3_ens"), row.get("joueur_carte3_val")),
            "bc1":  fmt_card(row.get("banquier_carte1_rang"), row.get("banquier_carte1_ens"), row.get("banquier_carte1_val")),
            "bc2":  fmt_card(row.get("banquier_carte2_rang"), row.get("banquier_carte2_ens"), row.get("banquier_carte2_val")),
            "bc3":  fmt_card(row.get("banquier_carte3_rang"), row.get("banquier_carte3_ens"), row.get("banquier_carte3_val")),
            "vict_j":   "\u2713" if gagnant == "Joueur"   else "",
            "vict_b":   "\u2713" if gagnant == "Banquier" else "",
            "egalite":  "Nul"    if gagnant == "Egalite"  else "",
            "nb_cartes": f"{pj_nb}/{pb_nb}",
            "pi_num":   "Pair" if num_jeu % 2 == 0 else "Impair",
            "pts_j":    str(pts_j),
            "sup65_j":  "\u2713" if pts_j >= 7 else "",
            "inf45_j":  "\u2713" if pts_j <= 4 else "",
            "pts_b":    str(pts_b),
            "sup65_b":  "\u2713" if pts_b >= 7 else "",
            "inf45_b":  "\u2713" if pts_b <= 4 else "",
        }

        pdf.set_fill_color(*bg)
        pdf.set_x(ML)
        for _, _, _, w, key in COLS:
            val = data.get(key, "")
            if key == "vict_j" and val:
                pdf.set_text_color(*C_BLUE); dv(True, 8)
            elif key == "vict_b" and val:
                pdf.set_text_color(*C_RED);  dv(True, 8)
            elif key == "egalite" and val:
                pdf.set_text_color(*C_GOLD); dv(True, 6)
            elif key == "pi_num":
                pdf.set_text_color(*(C_BLUE if val == "Pair" else C_RED)); dv(True, 5)
            elif key in ("sup65_j", "inf45_j", "sup65_b", "inf45_b") and val:
                pdf.set_text_color(*C_GREEN); dv(True, 8)
            else:
                pdf.set_text_color(*C_DARK); dv(False, 6)
            pdf.cell(w, ROW_H, val, border=1, fill=True, align='C')
        pdf.ln()

    def new_page_with_headers():
        pdf.add_page()
        draw_page_header()
        pdf.set_y(28)
        draw_col_headers()

    new_page_with_headers()

    for idx, row in enumerate(rows):
        if pdf.get_y() + ROW_H > pdf.h - 14:
            new_page_with_headers()
        draw_data_row(idx, row)

    if not rows:
        dv(False, 10)
        pdf.set_text_color(*C_GRAY)
        pdf.cell(0, 10, "Aucune partie enregistree pour cette date.", align='C',
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Pied de page
    pdf.set_y(pdf.h - 12)
    pdf.set_fill_color(*C_DARK)
    pdf.rect(0, pdf.h - 12, 297, 12, 'F')
    pdf.set_text_color(*C_WHITE)
    dv(False, 6)
    pdf.set_x(5)
    pdf.cell(287, 12,
             "(val) = valeur baccarat de la carte  |  \u2713 = Oui  |  Pair/Impair = parite du numero  |  "
             "Plus de 6.5 = total \u2265 7  |  Moins de 4.5 = total \u2264 4",
             align='C')

    pdf.output(path)


# ============================================================================
# COMMANDE /historique
# ============================================================================

async def cmd_start(event):
    if event.is_group or event.is_channel:
        return
    await event.respond(
        "👋 **Bienvenue sur le Bot Baccarat AI**\n\n"
        "Voici les commandes disponibles :\n\n"
        "/historique — 📅 Historique d'une date en PDF\n"
        "/recherche   — 🔍 Rechercher par filtre (couleurs, valeurs, seuils…)\n"
        "/comparaison — 📊 Comparer deux journées et trouver les communs\n"
        "/comptage   — 📊 Jeux enregistrés aujourd'hui\n"
        "/total      — 🗄️ Total global depuis la création de la base\n"
        "/taille     — 📦 Taille de la base de données"
    )


async def _process_historique(event, target_date_obj, target_date_str):
    """Génère et envoie le PDF pour une date donnée."""
    sender_id = event.sender_id
    JOURS = ["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
    jour_str = JOURS[target_date_obj.weekday()]

    msg = await event.respond(
        f"Recherche des parties du **{jour_str} {target_date_str}** en base de donnees..."
    )
    pdf_path = f"/tmp/hist_{sender_id}_{target_date_str.replace('/','')}.pdf"
    try:
        loop = asyncio.get_event_loop()
        count = await loop.run_in_executor(None, count_games_for_date, target_date_obj)
        await client.edit_message(
            event.chat_id, msg.id,
            f"{count} parties trouvees pour le {target_date_str}. Generation du PDF..."
        )
        rows  = await loop.run_in_executor(None, get_games_by_date, target_date_obj)
        stats = await loop.run_in_executor(None, get_stats_by_date, target_date_obj)
        await loop.run_in_executor(
            None, generate_pdf_from_db, pdf_path, rows, target_date_str, stats
        )
        total = int(stats.get("total", 0) or 0)
        jw    = int(stats.get("joueur_wins", 0) or 0)
        bw    = int(stats.get("banquier_wins", 0) or 0)
        egal  = int(stats.get("egalites", 0) or 0)
        nat   = int(stats.get("naturels", 0) or 0)
        pj    = int(stats.get("paires_joueur", 0) or 0)
        pb    = int(stats.get("paires_banquier", 0) or 0)

        def pct(v, t):
            return f"{round(v/t*100,1)}%" if t > 0 else "0%"

        caption = (
            f"📄 **Historique — {target_date_str}**\n"
            f"🕐 {get_local_time().strftime('%d/%m/%Y %H:%M')} GMT\n\n"
            f"🎮 **{total} parties** enregistrees\n"
            f"👤 Joueur  : {jw} victoires ({pct(jw,total)})\n"
            f"🏦 Banquier: {bw} victoires ({pct(bw,total)})\n"
            f"🤝 Egalite : {egal} ({pct(egal,total)})\n"
            f"⚡ Naturels: {nat} ({pct(nat,total)})\n"
            f"🃏 Paires J: {pj} | Paires B: {pb}\n\n"
            f"📋 Page 1 : tableau jeu par jeu (#{stats.get('premier_jeu','?')} → #{stats.get('dernier_jeu','?')})\n"
            f"📊 Page 2 : statistiques completes et distributions"
        )
        await client.delete_messages(event.chat_id, [msg.id])
        await client.send_file(
            event.chat_id, pdf_path,
            caption=caption, force_document=True, attributes=[],
        )
        logger.info(f"PDF historique {target_date_str} envoye a {sender_id} ({total} parties)")
    except Exception as e:
        logger.error(f"Erreur historique PDF: {e}")
        import traceback; traceback.print_exc()
        try:
            await client.edit_message(event.chat_id, msg.id,
                f"Erreur lors de la generation du PDF : {e}")
        except Exception:
            await event.respond(f"Erreur : {e}")
    finally:
        try:
            os.remove(pdf_path)
        except Exception:
            pass


async def cmd_historique(event):
    if event.is_group or event.is_channel:
        return
    if not is_admin(event.sender_id):
        await event.respond("Acces admin uniquement.")
        return

    sender_id = event.sender_id

    # Empêcher deux conversations simultanées
    if sender_id in active_conversations:
        await event.respond("Une demande est deja en cours. Patientez ou envoyez la date.")
        return
    active_conversations.add(sender_id)

    try:
        loop = asyncio.get_event_loop()
        dates = await loop.run_in_executor(None, get_dates_available)

        if dates:
            dates_str = "\n".join(
                f"  • {d.strftime('%d/%m/%Y')}"
                + (" — (aujourd'hui)" if d == current_game_date() else "")
                for d in dates[:10]
            )
            prompt = (
                "📅 **HISTORIQUE DES JEUX — Baccarat 1xBet**\n\n"
                "Dates disponibles en base de donnees :\n"
                f"{dates_str}\n\n"
                "Entrez la date souhaitee au format **JJ/MM/AAAA** :"
            )
        else:
            prompt = (
                "📅 **HISTORIQUE DES JEUX — Baccarat 1xBet**\n\n"
                "Aucune date disponible pour l'instant.\n"
                "Entrez quand meme une date au format **JJ/MM/AAAA** :"
            )

        # Ouvrir la conversation AVANT d'envoyer le prompt
        # pour que conv.get_response() sache quel message attendre
        async with client.conversation(event.chat_id, timeout=120, exclusive=False) as conv:
            await conv.send_message(prompt)

            target_date_obj = None
            target_date_str = None
            while True:
                try:
                    response = await conv.get_response(timeout=120)
                except asyncio.TimeoutError:
                    await client.send_message(
                        event.chat_id,
                        "⏱ Temps écoulé (2 min). Relancez /historique."
                    )
                    return

                text = response.message.strip()

                if text.startswith('/'):
                    await client.send_message(
                        event.chat_id,
                        "Saisie annulée. Relancez /historique si besoin."
                    )
                    return

                try:
                    parts = text.split('/')
                    if len(parts) != 3:
                        raise ValueError
                    day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                    target_date_obj = date_cls(year, month, day)
                    target_date_str = f"{day:02d}/{month:02d}/{year}"
                    break
                except ValueError:
                    await conv.send_message(
                        "Format invalide. Utilisez **JJ/MM/AAAA**\nEx : `23/03/2026`"
                    )

        if target_date_obj:
            await _process_historique(event, target_date_obj, target_date_str)

    finally:
        active_conversations.discard(sender_id)


# ============================================================================
# COMMANDE /comptage
# ============================================================================

async def cmd_comptage(event):
    """Jeux enregistrés aujourd'hui du #1 à l'heure actuelle."""
    if event.is_group or event.is_channel:
        return
    if not is_admin(event.sender_id):
        return

    now = get_local_time()
    today = now.date()

    # Numéro du jeu actuel (basé sur l'heure: 1 jeu/minute depuis 00:00)
    current_num = now.hour * 60 + now.minute + 1
    current_num = min(current_num, 1440)

    msg = await event.respond("Calcul en cours...")

    try:
        loop = asyncio.get_event_loop()
        data = await loop.run_in_executor(None, get_comptage_today, today, current_num)

        enreg    = int(data.get("enregistres", 0) or 0)
        dans_plage = int(data.get("dans_plage", 0) or 0)
        manquants  = int(data.get("manquants", 0) or 0)
        j_wins   = int(data.get("joueur_wins", 0) or 0)
        b_wins   = int(data.get("banquier_wins", 0) or 0)
        egal     = int(data.get("egalites", 0) or 0)
        premier  = data.get("premier") or "-"
        dernier  = data.get("dernier") or "-"

        def pct(v, t):
            return f"{round(v/t*100,1)}%" if t > 0 else "0%"

        couverture = pct(dans_plage, current_num)

        texte = (
            f"📊 **COMPTAGE DU JOUR — {today.strftime('%d/%m/%Y')}**\n"
            f"🕐 Heure actuelle : {now.strftime('%H:%M')} GMT\n\n"
            f"🎮 Jeu actuel estimé : **#{current_num}** / 1440\n"
            f"💾 Enregistrés dans la plage #1→#{current_num} : **{dans_plage}**\n"
            f"📈 Couverture : **{couverture}**\n"
            f"❌ Manquants (non capturés) : **{manquants}**\n\n"
            f"— Premier jeu en base : **#{premier}**\n"
            f"— Dernier jeu en base  : **#{dernier}**\n\n"
            f"🏆 Résultats des {enreg} parties enregistrées :\n"
            f"  👤 Joueur  : {j_wins} ({pct(j_wins, enreg)})\n"
            f"  🏦 Banquier: {b_wins} ({pct(b_wins, enreg)})\n"
            f"  🤝 Egalite : {egal} ({pct(egal, enreg)})\n\n"
            f"ℹ️ Le bot capture depuis son démarrage. "
            f"Les jeux manquants correspondent aux parties survenues avant le lancement du bot."
        )
        await client.edit_message(event.chat_id, msg.id, texte)
    except Exception as e:
        await client.edit_message(event.chat_id, msg.id, f"Erreur : {e}")


# ============================================================================
# COMMANDE /total
# ============================================================================

async def cmd_total(event):
    """Total global de toutes les parties en base depuis la création."""
    if event.is_group or event.is_channel:
        return
    if not is_admin(event.sender_id):
        return

    msg = await event.respond("Calcul du total global...")

    try:
        loop = asyncio.get_event_loop()
        data = await loop.run_in_executor(None, get_global_total)

        total  = int(data.get("total_global", 0) or 0)
        jours  = int(data.get("nb_jours", 0) or 0)
        pj     = data.get("premier_jour")
        dj     = data.get("dernier_jour")
        tj     = int(data.get("total_joueur", 0) or 0)
        tb     = int(data.get("total_banquier", 0) or 0)
        te     = int(data.get("total_egalite", 0) or 0)
        nat    = int(data.get("total_naturels", 0) or 0)
        j4c    = int(data.get("jeux_4c", 0) or 0)
        j5c    = int(data.get("jeux_5c", 0) or 0)
        j6c    = int(data.get("jeux_6c", 0) or 0)
        detail = data.get("detail_jours", [])

        def pct(v, t):
            return f"{round(v/t*100,1)}%" if t > 0 else "0%"

        # Tableau par jour
        lignes_jours = ""
        for d in detail:
            date_str = d["date_jeu"].strftime("%d/%m/%Y") if d.get("date_jeu") else "?"
            nb  = int(d.get("nb", 0) or 0)
            jw  = int(d.get("j", 0) or 0)
            bw  = int(d.get("b", 0) or 0)
            eg  = int(d.get("e", 0) or 0)
            mn  = d.get("min_jeu", "?")
            mx  = d.get("max_jeu", "?")
            lignes_jours += f"  `{date_str}` {nb:4d} parties | J:{jw} B:{bw} E:{eg} | #{mn}→#{mx}\n"

        texte = (
            f"🗄️ **TOTAL GLOBAL — BASE DE DONNÉES**\n"
            f"📅 Du {pj.strftime('%d/%m/%Y') if pj else '?'} au {dj.strftime('%d/%m/%Y') if dj else '?'}\n\n"
            f"🎮 **{total} parties** enregistrées sur **{jours} jour(s)**\n\n"
            f"🏆 Résultats globaux :\n"
            f"  👤 Joueur  : {tj:,} ({pct(tj, total)})\n"
            f"  🏦 Banquier: {tb:,} ({pct(tb, total)})\n"
            f"  🤝 Egalite : {te:,} ({pct(te, total)})\n\n"
            f"⚡ Naturels (8/9 en 2 cartes) : {nat} ({pct(nat, total)})\n"
            f"🃏 Parties 4 cartes : {j4c} ({pct(j4c, total)})\n"
            f"🃏 Parties 5 cartes : {j5c} ({pct(j5c, total)})\n"
            f"🃏 Parties 6 cartes : {j6c} ({pct(j6c, total)})\n\n"
            f"📋 **Détail par jour** (14 derniers) :\n"
            f"{lignes_jours}"
        )
        await client.edit_message(event.chat_id, msg.id, texte)
    except Exception as e:
        await client.edit_message(event.chat_id, msg.id, f"Erreur : {e}")


async def cmd_taille(event):
    """Taille de la base de données et documentation."""
    if event.is_group or event.is_channel:
        return
    if not is_admin(event.sender_id):
        return

    msg = await event.respond("Analyse de la base de données en cours...")

    try:
        loop = asyncio.get_event_loop()
        data = await loop.run_in_executor(None, get_db_size)

        nom_base     = data.get("nom_base", "?")
        taille_db    = data.get("taille_totale_db", "?")
        taille_t     = data.get("taille_table_parties", "?")
        taille_d     = data.get("taille_donnees_parties", "?")
        taille_i     = data.get("taille_index_parties", "?")
        total        = int(data.get("total", 0) or 0)
        jours        = int(data.get("jours", 0) or 0)
        octets_db    = int(data.get("octets_db", 0) or 0)
        octets_p     = int(data.get("octets_parties", 0) or 0)
        tables       = data.get("tables", [])

        # Calcul pourcentage d'utilisation
        pct_t = f"{round(octets_p/octets_db*100,1)}%" if octets_db > 0 else "0%"

        lignes_tables = ""
        for t in tables:
            lignes_tables += f"  `{t['tablename']}` — {t['taille']}\n"

        texte = (
            f"🗄️ **TAILLE DE LA BASE DE DONNÉES**\n"
            f"Base : `{nom_base}` (Render.com PostgreSQL)\n\n"
            f"📦 **Taille totale DB** : `{taille_db}`\n\n"
            f"📊 **Table `parties`** :\n"
            f"  Total données : `{taille_d}`\n"
            f"  Index         : `{taille_i}`\n"
            f"  Total (data+index) : `{taille_t}` ({pct_t} de la DB)\n\n"
            f"📋 **Contenu** :\n"
            f"  Parties enregistrées : **{total:,}** lignes\n"
            f"  Jours couverts       : **{jours}** jour(s)\n\n"
            f"📁 **Toutes les tables** :\n"
            f"{lignes_tables}\n"
            f"📖 La documentation technique est stockée dans la table `documentation`.\n"
            f"Un autre bot peut lire :\n"
            f"`SELECT section, cle, valeur FROM documentation ORDER BY section, cle;`"
        )
        await client.edit_message(event.chat_id, msg.id, texte)
    except Exception as e:
        await client.edit_message(event.chat_id, msg.id, f"Erreur : {e}")


FILTRES_LABELS = {
    # Config cartes
    "2/2": "2J + 2B (4 cartes)",    "3/2": "3J + 2B (5 cartes)",
    "2/3": "2J + 3B (5 cartes)",    "3/3": "3J + 3B (6 cartes)",
    # Résultat
    "joueur":   "Victoire Joueur",  "banquier": "Victoire Banquier",
    "egalite":  "Egalite (Nul)",
    # Événements
    "naturel":        "Naturels (8 ou 9 en 2 cartes)",
    "paire_joueur":   "Paire Joueur",
    "paire_banquier": "Paire Banquier",
    # Seuils Joueur
    "plus65_joueur":  "Joueur Plus de 6.5 (pts >= 7)",
    "moins45_joueur": "Joueur Moins de 4.5 (pts <= 4)",
    # Seuils Banquier
    "plus65_banquier":  "Banquier Plus de 6.5 (pts >= 7)",
    "moins45_banquier": "Banquier Moins de 4.5 (pts <= 4)",
    # Couleurs Joueur
    "pique_joueur":   "Joueur avec ♠ Pique",
    "trefle_joueur":  "Joueur avec ♣ Trefle",
    "carreau_joueur": "Joueur avec ♦ Carreau",
    "coeur_joueur":   "Joueur avec ♥ Coeur",
    # Couleurs Banquier
    "pique_banquier":   "Banquier avec ♠ Pique",
    "trefle_banquier":  "Banquier avec ♣ Trefle",
    "carreau_banquier": "Banquier avec ♦ Carreau",
    "coeur_banquier":   "Banquier avec ♥ Coeur",
    # Cartes hautes Joueur
    "as_joueur":    "Joueur avec As (A)",
    "roi_joueur":   "Joueur avec Roi (K)",
    "dame_joueur":  "Joueur avec Dame (Q)",
    "valet_joueur": "Joueur avec Valet (J)",
    "dix_joueur":   "Joueur avec Dix (10)",
    # Cartes hautes Banquier
    "as_banquier":    "Banquier avec As (A)",
    "roi_banquier":   "Banquier avec Roi (K)",
    "dame_banquier":  "Banquier avec Dame (Q)",
    "valet_banquier": "Banquier avec Valet (J)",
    "dix_banquier":   "Banquier avec Dix (10)",
}

FILTRES_SHORT = {
    "2/2": "2/2",    "3/2": "3/2",    "2/3": "2/3",    "3/3": "3/3",
    "joueur": "Joueur",  "banquier": "Banquier",  "egalite": "Egal.",
    "naturel": "Naturel", "paire_joueur": "Paire Joueur", "paire_banquier": "Paire Banquier",
    "plus65_joueur":   "+6.5 Joueur",   "moins45_joueur":   "-4.5 Joueur",
    "plus65_banquier": "+6.5 Banquier", "moins45_banquier": "-4.5 Banquier",
    "pique_joueur":   "♠ Joueur",   "trefle_joueur":   "♣ Joueur",   "carreau_joueur":   "♦ Joueur",   "coeur_joueur":   "♥ Joueur",
    "pique_banquier": "♠ Banquier", "trefle_banquier": "♣ Banquier", "carreau_banquier": "♦ Banquier", "coeur_banquier": "♥ Banquier",
    "as_joueur":    "As Joueur",    "roi_joueur":    "Roi Joueur",    "dame_joueur":    "Dame Joueur",    "valet_joueur":    "Valet Joueur",    "dix_joueur":    "Dix Joueur",
    "as_banquier":  "As Banquier",  "roi_banquier":  "Roi Banquier",  "dame_banquier":  "Dame Banquier",  "valet_banquier":  "Valet Banquier",  "dix_banquier":  "Dix Banquier",
}

FILTRES_JOUEUR = {
    "joueur", "egalite", "naturel", "paire_joueur",
    "plus65_joueur", "moins45_joueur",
    "pique_joueur", "trefle_joueur", "carreau_joueur", "coeur_joueur",
    "as_joueur", "roi_joueur", "dame_joueur", "valet_joueur", "dix_joueur",
    "2/2", "3/2", "2/3", "3/3",
}

FILTRES_BANQUIER = {
    "banquier", "egalite", "naturel", "paire_banquier",
    "plus65_banquier", "moins45_banquier",
    "pique_banquier", "trefle_banquier", "carreau_banquier", "coeur_banquier",
    "as_banquier", "roi_banquier", "dame_banquier", "valet_banquier", "dix_banquier",
    "2/2", "3/2", "2/3", "3/3",
}

FILTRES_MENU_JOUEUR = (
    "**Résultat**      : `joueur`  `egalite`\n"
    "**Évènements**    : `naturel`  `paire_joueur`\n"
    "**Seuils**        : `plus65_joueur`  `moins45_joueur`\n"
    "**Couleur**       : `pique_joueur`  `trefle_joueur`  `carreau_joueur`  `coeur_joueur`\n"
    "**Carte**         : `as_joueur`  `roi_joueur`  `dame_joueur`  `valet_joueur`  `dix_joueur`\n"
    "**Config cartes** : `2/2`  `3/2`  `2/3`  `3/3`"
)

FILTRES_MENU_BANQUIER = (
    "**Résultat**      : `banquier`  `egalite`\n"
    "**Évènements**    : `naturel`  `paire_banquier`\n"
    "**Seuils**        : `plus65_banquier`  `moins45_banquier`\n"
    "**Couleur**       : `pique_banquier`  `trefle_banquier`  `carreau_banquier`  `coeur_banquier`\n"
    "**Carte**         : `as_banquier`  `roi_banquier`  `dame_banquier`  `valet_banquier`  `dix_banquier`\n"
    "**Config cartes** : `2/2`  `3/2`  `2/3`  `3/3`"
)


async def _process_recherche(event, target_date_obj, target_date_str, filtres: list):
    """Génère et envoie le PDF multi-filtre (OR). filtres = liste de clés FILTRES_DISPONIBLES."""
    sender_id = event.sender_id

    # Labels et titre
    short_labels = [FILTRES_SHORT.get(f, f.upper()) for f in filtres]
    title_label  = " + ".join(short_labels)
    full_labels  = " | ".join(FILTRES_LABELS.get(f, f.upper()) for f in filtres)
    slug         = "_".join(f.replace("/","") for f in filtres)[:40]
    pdf_path     = f"/tmp/rech_{sender_id}_{target_date_str.replace('/','')}{slug}.pdf"

    msg = await event.respond(
        f"Recherche **[{title_label}]** pour le **{target_date_str}**..."
    )
    try:
        loop = asyncio.get_event_loop()

        # Recherche multi-filtre
        rows = await loop.run_in_executor(
            None, search_games_multi_filter, target_date_obj, filtres
        )
        if not rows:
            await client.edit_message(
                event.chat_id, msg.id,
                f"Aucune partie trouvée pour le **{target_date_str}** avec les filtres **{title_label}**."
            )
            return

        stats = stats_from_rows(rows)
        total = stats["total"]

        await client.edit_message(event.chat_id, msg.id,
            f"{total} partie(s) trouvée(s). Génération du PDF...")

        await loop.run_in_executor(
            None, generate_pdf_from_db, pdf_path, rows, target_date_str, stats, title_label
        )

        jw = stats["joueur_wins"]
        bw = stats["banquier_wins"]
        eg = stats["egalites"]
        def pct(v, t): return f"{round(v/t*100,1)}%" if t > 0 else "0%"

        # ── Liste taguée : Numéro X : ♣J + 2/2 ──────────────────────────────
        tagged_lines = ""
        for r in rows:
            num   = r.get("numero_jeu", "?")
            tags  = r.get("matched_filters", filtres)
            tag_s = " + ".join(FILTRES_SHORT.get(t, t) for t in tags)
            tagged_lines += f"  #{num} : {tag_s}\n"

        # Tronquer si trop long pour Telegram (max ~4096 chars)
        caption_base = (
            f"🔍 **RECHERCHE — {target_date_str}**\n"
            f"Filtre(s) : **{title_label}**\n"
            f"🕐 {get_local_time().strftime('%d/%m/%Y %H:%M')} GMT\n\n"
            f"🎮 **{total} partie(s)** trouvée(s)\n"
            f"👤 Joueur   : {jw} ({pct(jw,total)})\n"
            f"🏦 Banquier : {bw} ({pct(bw,total)})\n"
            f"🤝 Egalite  : {eg} ({pct(eg,total)})\n\n"
            f"📋 **Numéros trouvés** ({total}) :\n"
        )
        caption = caption_base + tagged_lines
        if len(caption) > 4000:
            # Tronquer la liste à ce qui rentre
            truncated = ""
            for line in tagged_lines.split("\n"):
                if len(caption_base) + len(truncated) + len(line) + 30 > 4000:
                    break
                truncated += line + "\n"
            caption = caption_base + truncated + f"  ... et {total - truncated.count('#')} autres"

        await client.delete_messages(event.chat_id, [msg.id])
        await client.send_file(
            event.chat_id, pdf_path,
            caption=caption, force_document=True, attributes=[],
        )
        logger.info(f"PDF recherche [{title_label}] {target_date_str} → {total} parties → {sender_id}")

    except Exception as e:
        logger.error(f"Erreur recherche PDF: {e}")
        import traceback; traceback.print_exc()
        try:
            await client.edit_message(event.chat_id, msg.id, f"Erreur : {e}")
        except Exception:
            await event.respond(f"Erreur : {e}")
    finally:
        try:
            os.remove(pdf_path)
        except Exception:
            pass


async def cmd_recherche(event):
    """Recherche des parties — 3 étapes : Date → Joueur/Banquier → Filtre(s)."""
    if event.is_group or event.is_channel:
        return
    if not is_admin(event.sender_id):
        await event.respond("Acces admin uniquement.")
        return

    sender_id = event.sender_id
    if sender_id in active_conversations:
        await event.respond("Une demande est deja en cours. Patientez.")
        return
    active_conversations.add(sender_id)

    try:
        loop = asyncio.get_event_loop()
        dates = await loop.run_in_executor(None, get_dates_available)

        dates_str = "\n".join(
            f"  • {d.strftime('%d/%m/%Y')}"
            + (" — (aujourd'hui)" if d == current_game_date() else "")
            for d in dates[:10]
        ) if dates else "  (aucune date disponible)"

        async with client.conversation(event.chat_id, timeout=120, exclusive=False) as conv:

            # ── ÉTAPE 1 : Date ────────────────────────────────────────────────
            await conv.send_message(
                "🔍 **RECHERCHE — Baccarat 1xBet**\n\n"
                f"Dates disponibles :\n{dates_str}\n\n"
                "**Étape 1/3** — Entrez la date **JJ/MM/AAAA** :"
            )
            target_date_obj = target_date_str = None
            while True:
                try:
                    resp = await conv.get_response(timeout=120)
                except asyncio.TimeoutError:
                    await client.send_message(event.chat_id, "⏱ Temps écoulé. Relancez /recherche.")
                    return
                text = resp.message.strip()
                if text.startswith('/'):
                    await client.send_message(event.chat_id, "Annulé.")
                    return
                try:
                    parts = text.split('/')
                    if len(parts) != 3:
                        raise ValueError
                    d2, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    target_date_obj = date_cls(y, m, d2)
                    target_date_str = f"{d2:02d}/{m:02d}/{y}"
                    break
                except ValueError:
                    await conv.send_message("Format invalide. Utilisez **JJ/MM/AAAA** — ex : `24/03/2026`")

            # ── ÉTAPE 2 : Côté (Joueur ou Banquier) ──────────────────────────
            await conv.send_message(
                f"Date : **{target_date_str}** ✓\n\n"
                "**Étape 2/3** — Choisissez le côté à analyser :\n\n"
                "`joueur` — Filtres côté Joueur\n"
                "`banquier` — Filtres côté Banquier"
            )
            cote = None
            filtres_set = None
            filtres_menu = None
            while True:
                try:
                    resp = await conv.get_response(timeout=120)
                except asyncio.TimeoutError:
                    await client.send_message(event.chat_id, "⏱ Temps écoulé. Relancez /recherche.")
                    return
                text = resp.message.strip().lower()
                if text.startswith('/'):
                    await client.send_message(event.chat_id, "Annulé.")
                    return
                if text in ("joueur", "j"):
                    cote, filtres_set, filtres_menu = "Joueur", FILTRES_JOUEUR, FILTRES_MENU_JOUEUR
                    break
                elif text in ("banquier", "b"):
                    cote, filtres_set, filtres_menu = "Banquier", FILTRES_BANQUIER, FILTRES_MENU_BANQUIER
                    break
                else:
                    await conv.send_message("Tapez `joueur` ou `banquier`.")

            # ── ÉTAPE 3 : Filtre(s) ──────────────────────────────────────────
            await conv.send_message(
                f"Côté : **{cote}** ✓\n\n"
                f"**Étape 3/3** — Filtres **{cote}** disponibles :\n{filtres_menu}\n\n"
                "Tapez **un ou plusieurs filtres** séparés par des virgules :\n"
                "ex : `trefle_joueur`  ou  `2/2, paire_joueur, naturel`"
            )
            filtres_choisis = None
            while True:
                try:
                    resp = await conv.get_response(timeout=120)
                except asyncio.TimeoutError:
                    await client.send_message(event.chat_id, "⏱ Temps écoulé. Relancez /recherche.")
                    return
                text = resp.message.strip()
                if text.startswith('/'):
                    await client.send_message(event.chat_id, "Annulé.")
                    return

                # Séparer par virgules ou espaces
                raw     = [p.strip().lower() for p in text.replace(",", " ").split() if p.strip()]
                valides = [p for p in raw if p in FILTRES_DISPONIBLES and p in filtres_set]
                hors    = [p for p in raw if p in FILTRES_DISPONIBLES and p not in filtres_set]
                inconnu = [p for p in raw if p not in FILTRES_DISPONIBLES]

                if not valides:
                    msg_err = "Aucun filtre reconnu pour ce côté. Réessayez."
                    if hors:
                        msg_err += f"\n(Filtre(s) de l'autre côté non applicables : `{'`, `'.join(hors)}`)"
                    await conv.send_message(msg_err)
                    continue

                avertissement = ""
                if hors:
                    avertissement += f"Filtre(s) de l'autre côté ignorés : `{'`, `'.join(hors)}`\n"
                if inconnu:
                    avertissement += f"Filtre(s) inconnus ignorés : `{'`, `'.join(inconnu)}`\n"
                if avertissement:
                    await conv.send_message(
                        avertissement + f"Filtres retenus : **{'**, **'.join(valides)}**\nTraitement en cours..."
                    )
                filtres_choisis = valides
                break

        if target_date_obj and filtres_choisis:
            await _process_recherche(event, target_date_obj, target_date_str, filtres_choisis)

    finally:
        active_conversations.discard(sender_id)


def generate_pdf_comparaison(path: str, data: list, date_a_str: str, date_b_str: str) -> None:
    """PDF paysage A4 : comparaison de deux journées côte à côte par numéro de jeu."""
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    FONT_TTF  = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
    FONT_TTFB = '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'

    C_DARK  = (20, 30, 60)
    C_WHITE = (255, 255, 255)
    C_HEAD  = (40, 60, 110)
    C_LIGHT = (235, 240, 255)
    C_ODD   = (250, 250, 255)
    C_GREEN = (20, 110, 50)
    C_RED   = (160, 20, 20)
    C_BLUE  = (20, 70, 180)
    C_GOLD  = (140, 100, 0)
    C_GRAY  = (130, 130, 130)
    C_MATCH = (0, 140, 60)    # vert foncé pour match
    C_DIFF  = (170, 30, 30)   # rouge pour différence

    pdf = FPDF(orientation='L', format='A4')
    pdf.add_font('DejaVu',  '',  FONT_TTF)
    pdf.add_font('DejaVu',  'B', FONT_TTFB)
    pdf.set_auto_page_break(auto=False)
    ML = 8
    now_str = get_local_time().strftime("%d/%m/%Y %H:%M")
    total = len(data)

    # Nombre de matchs totaux par catégorie
    def dv(bold=False, size=7):
        pdf.set_font('DejaVu', 'B' if bold else '', size)

    # ── Structure des colonnes ───────────────────────────────────────────────
    # Numéro | JOUR A (Heure Config Gagnant PtsJ PtsB) |
    # JOUR B (Heure Config Gagnant PtsJ PtsB) |
    # CORRESPONDANCES (Gagnant Config +6.5J -4.5J +6.5B -4.5B ♠J ♣J ♦J ♥J ♠B ♣B ♦B ♥B)

    COLS_A = [
        ("Heure",   11, "a_heure"),
        ("Config",  10, "a_config"),
        ("Gagnant", 15, "a_gagnant"),
        ("Pts J",    9, "a_pts_j"),
        ("Pts B",    9, "a_pts_b"),
        ("+6.5J",    9, "a_sup65_j"),
        ("-4.5J",    9, "a_inf45_j"),
        ("+6.5B",    9, "a_sup65_b"),
        ("-4.5B",    9, "a_inf45_b"),
        ("♠J",       7, "a_pique_j"),
        ("♣J",       7, "a_trefle_j"),
        ("♦J",       7, "a_carreau_j"),
        ("♥J",       7, "a_coeur_j"),
        ("♠B",       7, "a_pique_b"),
        ("♣B",       7, "a_trefle_b"),
        ("♦B",       7, "a_carreau_b"),
        ("♥B",       7, "a_coeur_b"),
    ]
    COLS_B = [(l, w, k.replace("a_", "b_")) for l, w, k in COLS_A]

    MATCH_COLS = [
        ("Gag.",   10, "match_gagnant"),
        ("Cfg.",    9, "match_config"),
        ("+6.5J",   9, "match_sup65_j"),
        ("-4.5J",   9, "match_inf45_j"),
        ("+6.5B",   9, "match_sup65_b"),
        ("-4.5B",   9, "match_inf45_b"),
        ("♠J",      7, "match_pique_j"),
        ("♣J",      7, "match_trefle_j"),
        ("♦J",      7, "match_carreau_j"),
        ("♥J",      7, "match_coeur_j"),
        ("♠B",      7, "match_pique_b"),
        ("♣B",      7, "match_trefle_b"),
        ("♦B",      7, "match_carreau_b"),
        ("♥B",      7, "match_coeur_b"),
    ]

    W_NUM   = 10
    W_A     = sum(w for _, w, _ in COLS_A)
    W_B     = sum(w for _, w, _ in COLS_B)
    W_MATCH = sum(w for _, w, _ in MATCH_COLS)
    W_TOTAL = W_NUM + W_A + W_B + W_MATCH

    ROW_H  = 6
    HDR1_H = 7
    HDR2_H = 10

    GRP_COLORS = {
        "num":   (130, 130, 160),
        "A":     (20, 70, 180),
        "B":     (160, 20, 20),
        "match": (20, 110, 50),
    }

    def draw_header():
        pdf.set_fill_color(*C_DARK); pdf.rect(0, 0, 297, 24, 'F')
        pdf.set_text_color(*C_WHITE)
        dv(True, 13)
        pdf.set_xy(ML, 2)
        pdf.cell(W_TOTAL, 9, f"COMPARAISON  {date_a_str}  vs  {date_b_str}", align='C',
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        dv(False, 7)
        pdf.set_xy(ML, 12)
        pdf.cell(W_TOTAL, 8,
                 f"{total} jeux communs  |  {now_str} GMT  |  "
                 f"Match gagnant : {sum(1 for r in data if r['match_gagnant'])}  |  "
                 f"Match config  : {sum(1 for r in data if r['match_config'])}",
                 align='C')

    def draw_col_headers():
        y = pdf.get_y()
        # Niveau 1 : groupes
        pdf.set_xy(ML, y)
        for grp, w, lbl in [
            ("num", W_NUM, ""), ("A", W_A, f"JOUR A — {date_a_str}"),
            ("B", W_B, f"JOUR B — {date_b_str}"), ("match", W_MATCH, "CORRESPONDANCES")
        ]:
            pdf.set_fill_color(*GRP_COLORS[grp])
            pdf.set_text_color(*C_WHITE); dv(True, 8)
            pdf.cell(w, HDR1_H, lbl, border=1, fill=True, align='C')
        pdf.ln()

        # Niveau 2 : colonnes
        y2 = pdf.get_y()
        x  = ML
        # Numéro
        pdf.set_fill_color(*C_HEAD); pdf.rect(x, y2, W_NUM, HDR2_H, 'FD')
        pdf.set_text_color(*C_WHITE); dv(True, 6)
        pdf.set_xy(x, y2 + 2); pdf.cell(W_NUM, 6, "N°", align='C')
        x += W_NUM
        for col_list in (COLS_A, COLS_B, MATCH_COLS):
            for lbl, w, _ in col_list:
                pdf.set_fill_color(*C_HEAD); pdf.rect(x, y2, w, HDR2_H, 'FD')
                pdf.set_text_color(*C_WHITE); dv(True, 5)
                pdf.set_xy(x, y2 + 2); pdf.cell(w, 6, lbl, align='C')
                x += w
        pdf.set_xy(ML, y2 + HDR2_H); pdf.ln()

    def draw_row(idx, row):
        bg = C_LIGHT if idx % 2 == 0 else C_ODD
        pdf.set_fill_color(*bg)
        pdf.set_x(ML)

        # Numéro
        pdf.set_text_color(*C_DARK); dv(True, 7)
        pdf.cell(W_NUM, ROW_H, str(row["numero_jeu"]), border=1, fill=True, align='C')

        # Jour A
        for lbl, w, key in COLS_A:
            val = row.get(key, "")
            if isinstance(val, bool):
                if val:
                    pdf.set_text_color(*C_GREEN); dv(True, 8)
                    pdf.cell(w, ROW_H, "\u2713", border=1, fill=True, align='C')
                else:
                    pdf.set_text_color(*C_GRAY); dv(False, 6)
                    pdf.cell(w, ROW_H, "", border=1, fill=True, align='C')
            else:
                if key == "a_gagnant":
                    col = C_BLUE if val == "Joueur" else (C_RED if val == "Banquier" else C_GOLD)
                    pdf.set_text_color(*col); dv(True, 6)
                else:
                    pdf.set_text_color(*C_DARK); dv(False, 6)
                pdf.cell(w, ROW_H, str(val), border=1, fill=True, align='C')

        # Jour B
        for lbl, w, key in COLS_B:
            val = row.get(key, "")
            if isinstance(val, bool):
                if val:
                    pdf.set_text_color(*C_GREEN); dv(True, 8)
                    pdf.cell(w, ROW_H, "\u2713", border=1, fill=True, align='C')
                else:
                    pdf.set_text_color(*C_GRAY); dv(False, 6)
                    pdf.cell(w, ROW_H, "", border=1, fill=True, align='C')
            else:
                if key == "b_gagnant":
                    col = C_BLUE if val == "Joueur" else (C_RED if val == "Banquier" else C_GOLD)
                    pdf.set_text_color(*col); dv(True, 6)
                else:
                    pdf.set_text_color(*C_DARK); dv(False, 6)
                pdf.cell(w, ROW_H, str(val), border=1, fill=True, align='C')

        # Correspondances
        for lbl, w, key in MATCH_COLS:
            val = row.get(key, False)
            if val:
                pdf.set_text_color(*C_MATCH); dv(True, 8)
                pdf.cell(w, ROW_H, "\u2713", border=1, fill=True, align='C')
            else:
                pdf.set_text_color(*C_DIFF); dv(True, 8)
                pdf.cell(w, ROW_H, "\u2717", border=1, fill=True, align='C')
        pdf.ln()

    def new_page():
        pdf.add_page()
        draw_header()
        pdf.set_y(26)
        draw_col_headers()

    new_page()
    for idx, row in enumerate(data):
        if pdf.get_y() + ROW_H > pdf.h - 12:
            new_page()
        draw_row(idx, row)

    if not data:
        dv(False, 10); pdf.set_text_color(*C_GRAY)
        pdf.cell(0, 10, "Aucun jeu commun entre ces deux dates.", align='C',
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Pied de page
    pdf.set_y(pdf.h - 10)
    pdf.set_fill_color(*C_DARK); pdf.rect(0, pdf.h - 10, 297, 10, 'F')
    pdf.set_text_color(*C_WHITE); dv(False, 6); pdf.set_x(5)
    nb_full = sum(1 for r in data if r["nb_matchs"] == len(MATCH_COLS))
    pdf.cell(287, 10,
             f"\u2713 = correspondance identique  |  \u2717 = difference  |  "
             f"Jeux avec toutes categories identiques : {nb_full}/{total}",
             align='C')
    pdf.output(path)


async def cmd_comparaison(event):
    """Compare deux journées et trouve les points communs par numéro de jeu."""
    if event.is_group or event.is_channel:
        return
    if not is_admin(event.sender_id):
        await event.respond("Acces admin uniquement.")
        return

    sender_id = event.sender_id
    if sender_id in active_conversations:
        await event.respond("Une demande est deja en cours. Patientez.")
        return
    active_conversations.add(sender_id)

    try:
        loop = asyncio.get_event_loop()
        dates = await loop.run_in_executor(None, get_dates_available)
        dates_str = "\n".join(
            f"  • {d.strftime('%d/%m/%Y')}"
            + (" — (aujourd'hui)" if d == current_game_date() else "")
            for d in dates[:10]
        ) if dates else "  (aucune date disponible)"

        def parse_date(text):
            parts = text.strip().split('/')
            if len(parts) != 3:
                raise ValueError
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            return date_cls(y, m, d), f"{d:02d}/{m:02d}/{y}"

        async with client.conversation(event.chat_id, timeout=120, exclusive=False) as conv:

            # ── DATE A ────────────────────────────────────────────────────────
            await conv.send_message(
                "📊 **COMPARAISON DE DEUX JOURNEES — Baccarat 1xBet**\n\n"
                f"Dates disponibles :\n{dates_str}\n\n"
                "**Étape 1/2** — Entrez la **première date** (JJ/MM/AAAA) :"
            )
            date_a_obj = date_a_str = None
            while True:
                try:
                    resp = await conv.get_response(timeout=120)
                except asyncio.TimeoutError:
                    await client.send_message(event.chat_id, "⏱ Temps écoulé. Relancez /comparaison.")
                    return
                if resp.message.strip().startswith('/'):
                    await client.send_message(event.chat_id, "Annulé.")
                    return
                try:
                    date_a_obj, date_a_str = parse_date(resp.message)
                    break
                except ValueError:
                    await conv.send_message("Format invalide. Utilisez **JJ/MM/AAAA** — ex : `23/03/2026`")

            # ── DATE B ────────────────────────────────────────────────────────
            await conv.send_message(
                f"Date A : **{date_a_str}** ✓\n\n"
                "**Étape 2/2** — Entrez la **deuxième date** à comparer (JJ/MM/AAAA) :"
            )
            date_b_obj = date_b_str = None
            while True:
                try:
                    resp = await conv.get_response(timeout=120)
                except asyncio.TimeoutError:
                    await client.send_message(event.chat_id, "⏱ Temps écoulé. Relancez /comparaison.")
                    return
                if resp.message.strip().startswith('/'):
                    await client.send_message(event.chat_id, "Annulé.")
                    return
                try:
                    date_b_obj, date_b_str = parse_date(resp.message)
                    if date_b_obj == date_a_obj:
                        await conv.send_message("Les deux dates doivent être différentes. Entrez une autre date :")
                        continue
                    break
                except ValueError:
                    await conv.send_message("Format invalide. Utilisez **JJ/MM/AAAA** — ex : `24/03/2026`")

        # ── Génération ────────────────────────────────────────────────────────
        pdf_path = f"/tmp/comp_{sender_id}_{date_a_str.replace('/','')}_vs_{date_b_str.replace('/','')}.pdf"
        msg = await event.respond(
            f"Comparaison **{date_a_str}** vs **{date_b_str}** en cours..."
        )
        try:
            comp_data = await loop.run_in_executor(None, compare_dates, date_a_obj, date_b_obj)
            total = len(comp_data)
            if total == 0:
                await client.edit_message(
                    event.chat_id, msg.id,
                    f"Aucun jeu commun entre le **{date_a_str}** et le **{date_b_str}**."
                )
                return

            nb_gag  = sum(1 for r in comp_data if r["match_gagnant"])
            nb_cfg  = sum(1 for r in comp_data if r["match_config"])
            nb_full = sum(1 for r in comp_data if r["nb_matchs"] == 14)

            await client.edit_message(
                event.chat_id, msg.id,
                f"{total} jeux communs trouvés. Génération du PDF..."
            )
            await loop.run_in_executor(
                None, generate_pdf_comparaison, pdf_path, comp_data, date_a_str, date_b_str
            )

            caption = (
                f"📊 **COMPARAISON — {date_a_str} vs {date_b_str}**\n"
                f"🕐 {get_local_time().strftime('%d/%m/%Y %H:%M')} GMT\n\n"
                f"🎮 **{total}** jeux communs (même numéro sur les 2 jours)\n\n"
                f"✅ **Correspondances exactes :**\n"
                f"  Même gagnant    : {nb_gag}/{total}\n"
                f"  Même config     : {nb_cfg}/{total}\n"
                f"  Toutes categ.   : {nb_full}/{total}\n\n"
                f"📋 Le PDF montre chaque jeu côte à côte avec ✓/✗ par catégorie\n"
                f"(Gagnant, Config, +6.5J/B, -4.5J/B, Couleurs ♠♣♦♥ J/B)"
            )
            await client.delete_messages(event.chat_id, [msg.id])
            await client.send_file(
                event.chat_id, pdf_path,
                caption=caption, force_document=True, attributes=[],
            )
            logger.info(f"PDF comparaison {date_a_str}/{date_b_str} → {total} jeux → {sender_id}")

            # ── Notification texte à l'administrateur ─────────────────────────
            notif_lines = []
            for r in comp_data:
                tags = []
                if r["match_gagnant"]:  tags.append(f"Gagnant ({r['a_gagnant']})")
                if r["match_config"]:   tags.append(f"Config ({r['a_config']})")
                if r["match_sup65_j"]:  tags.append("+6.5 Joueur")
                if r["match_inf45_j"]:  tags.append("-4.5 Joueur")
                if r["match_sup65_b"]:  tags.append("+6.5 Banquier")
                if r["match_inf45_b"]:  tags.append("-4.5 Banquier")
                suit_tags = []
                for suit, mkey in [("♠","pique"),("♣","trefle"),("♦","carreau"),("♥","coeur")]:
                    if r.get(f"match_{mkey}_j"): suit_tags.append(f"{suit}J")
                    if r.get(f"match_{mkey}_b"): suit_tags.append(f"{suit}B")
                if suit_tags: tags.append(" ".join(suit_tags))
                if len(tags) >= 3:   # Notifier seulement les jeux avec ≥3 correspondances
                    notif_lines.append(f"  #{r['numero_jeu']} : {' + '.join(tags)}")

            if notif_lines:
                notif_header = (
                    f"📊 **CORRESPONDANCES TROUVÉES — {date_a_str} vs {date_b_str}**\n"
                    f"🎮 {total} jeux communs\n"
                    f"✅ Même gagnant : {nb_gag}/{total} | Même config : {nb_cfg}/{total}\n\n"
                    f"🔥 **Jeux avec 3+ catégories identiques ({len(notif_lines)}) :**\n"
                )
                notif_msg = notif_header + "\n".join(notif_lines[:50])
                if len(notif_lines) > 50:
                    notif_msg += f"\n  ... et {len(notif_lines)-50} autres"
                try:
                    await client.send_message(ADMIN_ID, notif_msg)
                except Exception as ne:
                    logger.warning(f"Notif comparaison admin échouée: {ne}")

        except Exception as e:
            logger.error(f"Erreur comparaison PDF: {e}")
            import traceback; traceback.print_exc()
            try:
                await client.edit_message(event.chat_id, msg.id, f"Erreur : {e}")
            except Exception:
                await event.respond(f"Erreur : {e}")
        finally:
            try:
                os.remove(pdf_path)
            except Exception:
                pass

    finally:
        active_conversations.discard(sender_id)


# ============================================================================
# EXPORT EXCEL & SURVEILLANCE BASE DE DONNÉES
# ============================================================================

DB_LIMIT_BYTES   = 1_073_741_824          # 1 Go = limite Render.com Free Tier
DB_WARN_PCT      = 0.80                   # Alerte à 80 % (~858 Mo)
DB_CRITICAL_PCT  = 0.90                   # Critique à 90 % (~966 Mo)
DB_CHECK_SECONDS = 3600                   # Vérification toutes les heures

# Évite les doublons de notification (réinitialisé au redémarrage)
_db_alert_sent = {"warn": False, "critical": False}


def generate_excel_export(path: str, rows: list, generated_at: str, size_info: dict) -> int:
    """Génère un fichier Excel avec TOUTES les parties. Retourne le nombre de lignes."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Feuille principale : toutes les parties ──────────────────────────────
    ws = wb.active
    ws.title = "Toutes les parties"
    ws.sheet_view.showGridLines = True

    # Couleurs
    DARK_FILL  = PatternFill("solid", fgColor="141E3C")
    HEAD_FILL  = PatternFill("solid", fgColor="28406E")
    ODD_FILL   = PatternFill("solid", fgColor="EBF0FF")
    EVEN_FILL  = PatternFill("solid", fgColor="F5F7FF")
    WHITE_FONT = Font(color="FFFFFF", bold=True, size=9)
    DARK_FONT  = Font(color="141E3C", size=8)
    CTR        = Alignment(horizontal="center", vertical="center")
    THIN       = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ── Ligne de titre (fusion) ──────────────────────────────────────────────
    NCOLS = 30
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    tc = ws["A1"]
    tc.value  = f"BACCARAT 1xBet — Export complet  |  {generated_at} GMT  |  {len(rows)} parties  |  DB: {size_info.get('taille_totale_db','?')}"
    tc.fill   = DARK_FILL
    tc.font   = Font(color="FFFFFF", bold=True, size=11)
    tc.alignment = CTR
    ws.row_dimensions[1].height = 22

    # ── En-têtes de colonnes (ligne 2) ──────────────────────────────────────
    HEADERS = [
        "Date", "Numéro", "Heure",
        "Gagnant", "Naturel", "Paire J", "Paire B",
        # Joueur
        "J. Cartes", "J. Points",
        "J. C1 Rang", "J. C1 Coul.", "J. C1 Val.",
        "J. C2 Rang", "J. C2 Coul.", "J. C2 Val.",
        "J. C3 Rang", "J. C3 Coul.", "J. C3 Val.",
        # Banquier
        "B. Cartes", "B. Points",
        "B. C1 Rang", "B. C1 Coul.", "B. C1 Val.",
        "B. C2 Rang", "B. C2 Coul.", "B. C2 Val.",
        "B. C3 Rang", "B. C3 Coul.", "B. C3 Val.",
        # Total
        "Total Cartes",
    ]
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = HEAD_FILL; cell.font = WHITE_FONT
        cell.alignment = CTR; cell.border = THIN
    ws.row_dimensions[2].height = 16

    # ── Largeurs de colonnes ─────────────────────────────────────────────────
    WIDTHS = [12, 8, 8, 12, 8, 8, 8,
              8, 8, 8, 8, 7, 8, 8, 7, 8, 8, 7,
              8, 8, 8, 8, 7, 8, 8, 7, 8, 8, 7,
              10]
    for ci, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Données ──────────────────────────────────────────────────────────────
    for ri, r in enumerate(rows, 3):
        fill = ODD_FILL if ri % 2 == 1 else EVEN_FILL
        def v(key, default=""):
            val = r.get(key)
            return val if val is not None else default
        def fmt_d(d):
            return d.strftime("%d/%m/%Y") if d else ""
        def fmt_t(t):
            return str(t)[:5] if t else ""
        def bool_v(key):
            return "Oui" if r.get(key) else ""

        row_data = [
            fmt_d(r.get("date_jeu")), v("numero_jeu"), fmt_t(r.get("heure")),
            v("gagnant"), bool_v("naturel"), bool_v("est_paire_joueur"), bool_v("est_paire_banquier"),
            v("joueur_total_cartes"), v("joueur_points"),
            v("joueur_carte1_rang"), v("joueur_carte1_ens"), v("joueur_carte1_val"),
            v("joueur_carte2_rang"), v("joueur_carte2_ens"), v("joueur_carte2_val"),
            v("joueur_carte3_rang"), v("joueur_carte3_ens"), v("joueur_carte3_val"),
            v("banquier_total_cartes"), v("banquier_points"),
            v("banquier_carte1_rang"), v("banquier_carte1_ens"), v("banquier_carte1_val"),
            v("banquier_carte2_rang"), v("banquier_carte2_ens"), v("banquier_carte2_val"),
            v("banquier_carte3_rang"), v("banquier_carte3_ens"), v("banquier_carte3_val"),
            v("total_cartes"),
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.font = DARK_FONT
            cell.alignment = CTR; cell.border = THIN
        ws.row_dimensions[ri].height = 13

    # Figer la ligne d'en-tête
    ws.freeze_panes = "A3"

    # ── Feuille récapitulatif ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Récapitulatif")
    recap = [
        ("Généré le",             generated_at + " GMT"),
        ("Taille totale DB",      size_info.get("taille_totale_db", "?")),
        ("Taille table parties",  size_info.get("taille_table_parties", "?")),
        ("Limite Render.com",     "1 Go (1 073 741 824 octets)"),
        ("Occupation (%)",        f"{size_info.get('octets_db',0)/DB_LIMIT_BYTES*100:.1f} %"),
        ("Total parties",         len(rows)),
        ("Jours couverts",        size_info.get("jours", 0)),
    ]
    for ri2, (k, val) in enumerate(recap, 1):
        ws2.cell(row=ri2, column=1, value=k).font   = Font(bold=True)
        ws2.cell(row=ri2, column=2, value=str(val))
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 25

    wb.save(path)
    return len(rows)


async def send_db_alert(level: str, size_info: dict) -> None:
    """Envoie une alerte DB à l'admin et, si critique, le fichier Excel complet."""
    pct = size_info.get("octets_db", 0) / DB_LIMIT_BYTES * 100
    taille = size_info.get("taille_totale_db", "?")
    total  = size_info.get("total", 0)
    jours  = size_info.get("jours", 0)

    if level == "warn":
        msg = (
            "⚠️ **ALERTE BASE DE DONNÉES — 80 % utilisé**\n\n"
            f"📦 Taille actuelle  : **{taille}** ({pct:.1f} %)\n"
            f"🎮 Parties stockées : **{total}**\n"
            f"📅 Jours couverts   : **{jours}**\n"
            f"🔴 Limite Render    : 1 Go\n\n"
            "Vous approchez de la limite de votre base de données Render.com.\n"
            "Il est conseillé de créer une nouvelle base dès que possible\n"
            "et de mettre à jour la variable **RENDER_DB_URL** dans les secrets."
        )
        await client.send_message(ADMIN_ID, msg)

    elif level == "critical":
        msg = (
            "🚨 **ALERTE CRITIQUE — 90 % BASE DE DONNÉES PLEINE**\n\n"
            f"📦 Taille actuelle  : **{taille}** ({pct:.1f} %)\n"
            f"🎮 Parties stockées : **{total}**\n"
            f"📅 Jours couverts   : **{jours}**\n"
            f"🔴 Limite Render    : 1 Go\n\n"
            "⚠️ **ACTION URGENTE REQUISE :**\n"
            "1. Créez une nouvelle base de données sur Render.com\n"
            "2. Mettez à jour **RENDER_DB_URL** dans les secrets du bot\n"
            "3. Redémarrez le bot\n\n"
            "📊 Export complet de toutes vos données en cours d'envoi..."
        )
        await client.send_message(ADMIN_ID, msg)

        # Export Excel complet
        now_str  = get_local_time().strftime("%Y%m%d_%H%M")
        xls_path = f"/tmp/backup_baccarat_{now_str}.xlsx"
        try:
            loop = asyncio.get_event_loop()
            rows = await loop.run_in_executor(None, get_all_games_for_export)
            nb   = await loop.run_in_executor(
                None, generate_excel_export,
                xls_path, rows, get_local_time().strftime("%d/%m/%Y %H:%M"), size_info
            )
            caption = (
                f"📊 **Export complet Baccarat 1xBet**\n"
                f"📅 Généré le {get_local_time().strftime('%d/%m/%Y %H:%M')} GMT\n"
                f"🎮 {nb} parties — {jours} jours\n"
                f"📦 Taille DB : {taille} ({pct:.1f} %)\n\n"
                f"Sauvegardez ce fichier avant de changer de base de données."
            )
            await client.send_file(
                ADMIN_ID, xls_path,
                caption=caption, force_document=True, attributes=[]
            )
            logger.info(f"Export Excel critique envoyé → {nb} parties")
        except Exception as e:
            logger.error(f"Erreur export Excel alerte critique : {e}")
            await client.send_message(ADMIN_ID, f"Erreur lors de la génération du fichier Excel : {e}")
        finally:
            try: os.remove(xls_path)
            except Exception: pass


async def db_monitor_loop() -> None:
    """Surveille la taille de la base toutes les heures et alerte l'admin si nécessaire."""
    logger.info("🗄️ Surveillance taille DB démarrée (vérification toutes les heures)")
    while True:
        await asyncio.sleep(DB_CHECK_SECONDS)
        try:
            loop     = asyncio.get_event_loop()
            info     = await loop.run_in_executor(None, get_db_size)
            octets   = info.get("octets_db", 0)
            pct      = octets / DB_LIMIT_BYTES

            if pct >= DB_CRITICAL_PCT and not _db_alert_sent["critical"]:
                _db_alert_sent["critical"] = True
                _db_alert_sent["warn"]     = True   # warn déjà couvert
                logger.warning(f"DB CRITIQUE {pct*100:.1f}% — envoi alerte admin + Excel")
                await send_db_alert("critical", info)

            elif pct >= DB_WARN_PCT and not _db_alert_sent["warn"]:
                _db_alert_sent["warn"] = True
                logger.warning(f"DB ALERTE {pct*100:.1f}% — envoi notification admin")
                await send_db_alert("warn", info)

            else:
                logger.info(f"DB OK — {info.get('taille_totale_db','?')} ({pct*100:.1f}%)")

        except Exception as e:
            logger.error(f"db_monitor_loop error: {e}")


# ============================================================================
# SETUP HANDLERS
# ============================================================================

def setup_handlers():
    client.add_event_handler(cmd_start,       events.NewMessage(pattern=r'^/start$'))
    client.add_event_handler(cmd_historique,  events.NewMessage(pattern=r'^/historique$'))
    client.add_event_handler(cmd_comptage,    events.NewMessage(pattern=r'^/comptage$'))
    client.add_event_handler(cmd_total,       events.NewMessage(pattern=r'^/total$'))
    client.add_event_handler(cmd_taille,      events.NewMessage(pattern=r'^/taille$'))
    client.add_event_handler(cmd_recherche,   events.NewMessage(pattern=r'^/recherche$'))
    client.add_event_handler(cmd_comparaison, events.NewMessage(pattern=r'^/comparaison$'))


# ============================================================================
# DEMARRAGE
# ============================================================================

async def register_commands():
    """Enregistre les commandes dans le menu Telegram du bot."""
    try:
        await client(SetBotCommandsRequest(
            scope=BotCommandScopeDefault(),
            lang_code="fr",
            commands=[
                BotCommand(command="start",      description="🤖 Démarrer le bot / voir les commandes"),
                BotCommand(command="historique", description="📅 Historique d'une date en PDF"),
                BotCommand(command="comptage",   description="📊 Jeux enregistrés aujourd'hui du #1 à maintenant"),
                BotCommand(command="total",      description="🗄️ Total global depuis la création de la base"),
                BotCommand(command="taille",     description="📦 Taille de la base de données"),
                BotCommand(command="recherche",   description="🔍 Rechercher par filtre (2/2, couleur, valeur…)"),
                BotCommand(command="comparaison", description="📊 Comparer deux journées et trouver les communs"),
            ]
        ))
        logger.info("Commandes enregistrées dans Telegram")
    except Exception as e:
        logger.error(f"Erreur enregistrement commandes: {e}")


async def start_bot():
    global client
    # Si TELEGRAM_SESSION est fourni (StringSession), on l'utilise
    # Sinon on utilise une session fichier persistante pour éviter
    # les re-authentifications à chaque redémarrage
    if TELEGRAM_SESSION:
        session = StringSession(TELEGRAM_SESSION)
    else:
        session = "/home/runner/workspace/bot_session"  # fichier .session persistant
    client = TelegramClient(session, API_ID, API_HASH)
    try:
        await client.start(bot_token=BOT_TOKEN)
        setup_handlers()
        await register_commands()
        logger.info("Bot demarré avec succes")
        return True
    except FloodWaitError as e:
        wait = e.seconds + 5
        logger.warning(f"FloodWait Telegram : attente automatique de {wait} secondes ({wait//60}m{wait%60}s)...")
        await asyncio.sleep(wait)
        # Réessayer après l'attente
        try:
            await client.start(bot_token=BOT_TOKEN)
            setup_handlers()
            await register_commands()
            logger.info("Bot demarré avec succes (apres FloodWait)")
            return True
        except Exception as e2:
            logger.error(f"Erreur demarrage bot apres FloodWait: {e2}")
            return False
    except Exception as e:
        logger.error(f"Erreur demarrage bot: {e}")
        return False


async def main():
    try:
        # Init DB
        try:
            init_db()
        except Exception as e:
            logger.error(f"Erreur init DB: {e}")

        # Documentation dans la base (pour bots secondaires)
        try:
            init_documentation()
        except Exception as e:
            logger.error(f"Erreur init documentation: {e}")

        if not await start_bot():
            return

        # Serveur web health check
        app = web.Application()
        app.router.add_get('/health', lambda r: web.Response(text="OK"))
        app.router.add_get('/', lambda r: web.Response(text="Baccarat AI - Running"))
        runner = web.AppRunner(app)
        await runner.setup()
        site = web.TCPSite(runner, '0.0.0.0', PORT)
        await site.start()
        logger.info(f"Serveur web sur port {PORT}")

        # Lancer la boucle de capture en arrière-plan
        asyncio.ensure_future(polling_loop())
        # Surveiller la taille de la DB toutes les heures
        asyncio.ensure_future(db_monitor_loop())

        await client.run_until_disconnected()

    except Exception as e:
        logger.error(f"Erreur main: {e}")
    finally:
        if client and client.is_connected():
            await client.disconnect()


if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Arrete par l'utilisateur")
    except Exception as e:
        logger.error(f"Fatal: {e}")
        sys.exit(1)
